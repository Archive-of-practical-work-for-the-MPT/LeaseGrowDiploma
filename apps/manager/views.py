"""Представления для менеджера: статистика, экспорт, чат."""
import json
from decimal import Decimal

from django.db.models import Count, Sum
from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.views import View

from apps.leasing.models import Company, LeaseContract, PaymentSchedule, MaintenanceRequest
from apps.catalog.models import Equipment, EquipmentCategory
from apps.control_panel.mixins import AdminRequiredMixin, AdminOrManagerRequiredMixin
from .mixins import ManagerRequiredMixin


def _get_statistics_data():
    """Собирает данные для статистики."""
    contracts_by_status = (
        LeaseContract.objects.values('status')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    status_labels = {'draft': 'Черновик', 'active': 'Действует', 'completed': 'Завершён', 'terminated': 'Расторгнут'}
    pie_contracts = {
        'labels': [status_labels.get(c['status'], c['status']) for c in contracts_by_status],
        'data': [c['count'] for c in contracts_by_status],
    }

    equipment_by_category = (
        Equipment.objects.values('category__name')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )
    bar_equipment = {
        'labels': [e['category__name'] or 'Без категории' for e in equipment_by_category],
        'data': [e['count'] for e in equipment_by_category],
    }

    payments_by_status = (
        PaymentSchedule.objects.values('status')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    payment_labels = {'pending': 'Ожидает', 'paid': 'Оплачен', 'overdue': 'Просрочен', 'cancelled': 'Отменён'}
    pie_payments = {
        'labels': [payment_labels.get(p['status'], p['status']) for p in payments_by_status],
        'data': [p['count'] for p in payments_by_status],
    }

    maint_requests_by_status = (
        MaintenanceRequest.objects.values('status')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    maint_labels = {'new': 'Новая', 'in_progress': 'В работе', 'completed': 'Выполнена', 'cancelled': 'Отменена'}
    pie_maintenance = {
        'labels': [maint_labels.get(m['status'], m['status']) for m in maint_requests_by_status],
        'data': [m['count'] for m in maint_requests_by_status],
    }

    total_amount = LeaseContract.objects.filter(status='active').aggregate(s=Sum('total_amount'))['s'] or Decimal('0')
    contracts_count = LeaseContract.objects.count()
    companies_count = Company.objects.count()
    equipment_count = Equipment.objects.count()

    return {
        'pie_contracts': pie_contracts,
        'bar_equipment': bar_equipment,
        'pie_payments': pie_payments,
        'pie_maintenance': pie_maintenance,
        'total_amount': total_amount,
        'contracts_count': contracts_count,
        'companies_count': companies_count,
        'equipment_count': equipment_count,
    }


def _stats_with_json():
    """Статистика с JSON-строками для шаблона."""
    data = _get_statistics_data()
    data['pie_contracts_labels_json'] = json.dumps(data['pie_contracts']['labels'])
    data['pie_contracts_data_json'] = json.dumps(data['pie_contracts']['data'])
    data['pie_payments_labels_json'] = json.dumps(data['pie_payments']['labels'])
    data['pie_payments_data_json'] = json.dumps(data['pie_payments']['data'])
    data['pie_maintenance_labels_json'] = json.dumps(data['pie_maintenance']['labels'])
    data['pie_maintenance_data_json'] = json.dumps(data['pie_maintenance']['data'])
    data['bar_equipment_labels_json'] = json.dumps(data['bar_equipment']['labels'])
    data['bar_equipment_data_json'] = json.dumps(data['bar_equipment']['data'])
    return data


def _get_admin_stats():
    """Статистика системы для администратора."""
    from apps.accounts.models import Account, Role
    from apps.catalog.models import Equipment, EquipmentCategory, Manufacturer
    from apps.leasing.models import Company, LeaseContract, PaymentSchedule, Maintenance, MaintenanceRequest
    return {
        'accounts': Account.objects.count(),
        'roles': Role.objects.count(),
        'companies': Company.objects.count(),
        'equipment': Equipment.objects.count(),
        'categories': EquipmentCategory.objects.count(),
        'manufacturers': Manufacturer.objects.count(),
        'contracts': LeaseContract.objects.count(),
        'contracts_active': LeaseContract.objects.filter(status='active').count(),
        'payments': PaymentSchedule.objects.count(),
        'payments_pending': PaymentSchedule.objects.filter(status='pending').count(),
        'maintenance': Maintenance.objects.count(),
        'maintenance_requests': MaintenanceRequest.objects.count(),
        'maintenance_requests_new': MaintenanceRequest.objects.filter(status='new').count(),
    }


class StatisticsView(ManagerRequiredMixin, View):
    """Статистика с диаграммами и экспортом. Только для менеджера."""

    def get(self, request):
        data = _stats_with_json()
        return render(request, 'manager/statistics.html', data)


class AdminStatisticsView(AdminRequiredMixin, View):
    """Статистика системы (только карточки). Только для администратора."""

    def get(self, request):
        stats = _get_admin_stats()
        return render(request, 'manager/admin_statistics.html', {'stats': stats})


def _build_export_data():
    """Строит плоскую структуру для экспорта."""
    data = _get_statistics_data()
    rows = []

    rows.append(('Сводка', '', ''))
    rows.append(('Компаний', data['companies_count'], ''))
    rows.append(('Техники (единиц)', data['equipment_count'], ''))
    rows.append(('Договоров всего', data['contracts_count'], ''))
    rows.append(('Сумма активных договоров', str(data['total_amount']), 'руб.'))
    rows.append(('', '', ''))
    rows.append(('Договоры по статусам', 'Количество', ''))
    for label, val in zip(data['pie_contracts']['labels'], data['pie_contracts']['data']):
        rows.append((label, val, ''))
    rows.append(('', '', ''))
    rows.append(('Платежи по статусам', 'Количество', ''))
    for label, val in zip(data['pie_payments']['labels'], data['pie_payments']['data']):
        rows.append((label, val, ''))
    rows.append(('', '', ''))
    rows.append(('Заявки на обслуживание', 'Количество', ''))
    for label, val in zip(data['pie_maintenance']['labels'], data['pie_maintenance']['data']):
        rows.append((label, val, ''))
    rows.append(('', '', ''))
    rows.append(('Техника по категориям', 'Количество', ''))
    for label, val in zip(data['bar_equipment']['labels'], data['bar_equipment']['data']):
        rows.append((label, val, ''))

    return rows


class StatisticsExportExcelView(ManagerRequiredMixin, View):
    """Экспорт статистики в Excel."""

    def get(self, request):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse(
                'Для Excel-экспорта установите openpyxl: pip install openpyxl',
                status=500,
            )

        rows = _build_export_data()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Статистика LeaseGrow'

        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font_white = Font(bold=True, size=12, color='FFFFFF')

        ws.append(['Статистика LeaseGrow', '', ''])
        ws.merge_cells('A1:C1')
        ws['A1'].font = Font(bold=True, size=14)
        ws.append(['Дата формирования:', timezone.now().strftime('%d.%m.%Y %H:%M'), ''])
        ws.append([])

        for i, row in enumerate(rows, start=5):
            ws.append(row)
            if row[0] in ('Сводка', 'Договоры по статусам', 'Платежи по статусам',
                          'Заявки на обслуживание', 'Техника по категориям'):
                for col in ('A', 'B', 'C'):
                    cell = ws[f'{col}{i}']
                    cell.font = header_font
                    cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

        for idx, col in enumerate(ws.columns, start=1):
            col_letter = get_column_letter(idx)
            max_length = 0
            for cell in col:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        # Лист с диаграммами
        data = _get_statistics_data()
        ws_charts = wb.create_sheet('Диаграммы', 1)

        def add_pie_chart(ws_target, title, labels, values, row_start, col_chart='D'):
            """Добавляет круговую диаграмму."""
            ws_target.cell(row=row_start, column=1, value=title)
            for i, (lb, val) in enumerate(zip(labels, values), start=row_start + 1):
                ws_target.cell(row=i, column=1, value=lb)
                ws_target.cell(row=i, column=2, value=val)
            row_end = row_start + len(labels)
            from openpyxl.chart import PieChart, Reference
            pie = PieChart()
            pie.title = title
            data_ref = Reference(ws_target, min_col=2, min_row=row_start + 1, max_row=row_end)
            cats = Reference(ws_target, min_col=1, min_row=row_start + 1, max_row=row_end)
            pie.add_data(data_ref, titles_from_data=False)
            pie.set_categories(cats)
            ws_target.add_chart(pie, f'{col_chart}{row_start}')
            return row_end + 3

        def add_bar_chart(ws_target, title, labels, values, row_start, col_chart='D'):
            """Добавляет столбчатую диаграмму."""
            ws_target.cell(row=row_start, column=1, value='Категория')
            ws_target.cell(row=row_start, column=2, value='Количество')
            for i, (lb, val) in enumerate(zip(labels, values), start=row_start + 1):
                ws_target.cell(row=i, column=1, value=lb)
                ws_target.cell(row=i, column=2, value=val)
            row_end = row_start + len(labels)
            from openpyxl.chart import BarChart, Reference
            bar = BarChart()
            bar.type = 'col'
            bar.title = title
            data_ref = Reference(ws_target, min_col=2, min_row=row_start, max_row=row_end)
            cats = Reference(ws_target, min_col=1, min_row=row_start + 1, max_row=row_end)
            bar.add_data(data_ref, titles_from_data=True)
            bar.set_categories(cats)
            ws_target.add_chart(bar, f'{col_chart}{row_start}')
            return row_end + 3

        row = 1
        row = add_pie_chart(
            ws_charts, 'Договоры по статусам',
            data['pie_contracts']['labels'], data['pie_contracts']['data'], row
        )
        row = add_pie_chart(
            ws_charts, 'Платежи по статусам',
            data['pie_payments']['labels'], data['pie_payments']['data'], row
        )
        row = add_pie_chart(
            ws_charts, 'Заявки на обслуживание',
            data['pie_maintenance']['labels'], data['pie_maintenance']['data'], row
        )
        add_bar_chart(
            ws_charts, 'Техника по категориям',
            data['bar_equipment']['labels'], data['bar_equipment']['data'], row
        )

        filename = f'leasegrow_statistics_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class StatisticsExportPdfView(ManagerRequiredMixin, View):
    """Экспорт статистики в PDF."""

    def get(self, request):
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            return HttpResponse(
                'Для PDF-экспорта установите reportlab: pip install reportlab',
                status=500,
            )

        import os
        font_name = 'Helvetica'
        if os.name == 'nt':
            fonts_dir = os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts')
            arial_path = os.path.join(fonts_dir, 'arial.ttf')
            if os.path.exists(arial_path):
                try:
                    pdfmetrics.registerFont(TTFont('Arial', arial_path))
                    font_name = 'Arial'
                except Exception:
                    pass

        response = HttpResponse(content_type='application/pdf')
        filename = f'leasegrow_statistics_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        style_heading = ParagraphStyle(
            name='CustomHeading',
            parent=styles['Heading1'],
            fontName=font_name,
        )
        style_body = ParagraphStyle(
            name='CustomBody',
            parent=styles['Normal'],
            fontName=font_name,
        )

        story = []
        story.append(Paragraph('Статистика LeaseGrow', style_heading))
        story.append(Spacer(1, 12))
        story.append(Paragraph(f'Дата формирования: {timezone.now().strftime("%d.%m.%Y %H:%M")}', style_body))
        story.append(Spacer(1, 20))

        rows = _build_export_data()
        table_data = [[str(cell) for cell in row] for row in rows if any(cell for cell in row)]
        table = Table(table_data, colWidths=[8*cm, 4*cm, 3*cm])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E2EFDA')),
            ('FONTNAME', (0, 0), (-1, 0), font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
        ]))
        story.append(table)
        story.append(Spacer(1, 30))

        # Диаграммы в PDF через matplotlib
        try:
            import matplotlib
            matplotlib.use('Agg')
            import matplotlib.pyplot as plt
            plt.rcParams['font.family'] = 'DejaVu Sans'
            from io import BytesIO
            from reportlab.platypus import Image

            data = _get_statistics_data()
            figsize = (5, 3.5)
            colors = ['#3b82f6', '#10b981', '#f59e0b', '#ef4444', '#8b5cf6', '#ec4899']

            def make_pie_image(title, labels, values):
                if not values:
                    return None
                fig, ax = plt.subplots(figsize=figsize)
                ax.pie(values, labels=labels, autopct='%1.1f%%', colors=colors[:len(values)])
                ax.set_title(title, fontsize=10)
                buf = BytesIO()
                plt.tight_layout()
                plt.savefig(buf, format='png', dpi=120, bbox_inches='tight')
                plt.close()
                buf.seek(0)
                return Image(buf, width=14*cm, height=10*cm)

            def make_bar_image(title, labels, values):
                if not values:
                    return None
                fig, ax = plt.subplots(figsize=figsize)
                bars = ax.bar(range(len(labels)), values, color=colors[0])
                ax.set_xticks(range(len(labels)))
                ax.set_xticklabels(labels, rotation=45, ha='right', fontsize=8)
                ax.set_title(title, fontsize=10)
                ax.set_ylabel('Количество')
                buf = BytesIO()
                plt.tight_layout()
                plt.savefig(buf, format='png', dpi=120, bbox_inches='tight')
                plt.close()
                buf.seek(0)
                return Image(buf, width=14*cm, height=10*cm)

            story.append(Paragraph('Диаграммы', style_heading))
            story.append(Spacer(1, 12))

            img1 = make_pie_image('Договоры по статусам', data['pie_contracts']['labels'], data['pie_contracts']['data'])
            if img1:
                story.append(img1)
                story.append(Spacer(1, 15))
            img2 = make_pie_image('Платежи по статусам', data['pie_payments']['labels'], data['pie_payments']['data'])
            if img2:
                story.append(img2)
                story.append(Spacer(1, 15))
            img3 = make_pie_image('Заявки на обслуживание', data['pie_maintenance']['labels'], data['pie_maintenance']['data'])
            if img3:
                story.append(img3)
                story.append(Spacer(1, 15))
            img4 = make_bar_image('Техника по категориям', data['bar_equipment']['labels'], data['bar_equipment']['data'])
            if img4:
                story.append(img4)
        except ImportError:
            pass  # matplotlib не установлен — экспортируем без диаграмм

        doc.build(story)
        return response


class LeaseRequestCreateContractView(AdminOrManagerRequiredMixin, View):
    """Оформление договора из подтверждённой заявки."""

    def get(self, request, pk):
        from apps.leasing.models import LeaseRequest
        from apps.leasing.forms import ContractFromRequestForm
        lease_req = get_object_or_404(LeaseRequest, pk=pk)
        if lease_req.status != 'confirmed':
            messages.error(request, 'Можно оформить договор только по подтверждённой заявке.')
            return redirect('chat:thread', request_id=pk)
        form = ContractFromRequestForm(lease_request=lease_req)
        return render(request, 'manager/lease_request_create_contract.html', {
            'lease_request': lease_req,
            'form': form,
        })

    def post(self, request, pk):
        from apps.leasing.models import LeaseRequest
        from apps.leasing.forms import ContractFromRequestForm
        lease_req = get_object_or_404(LeaseRequest, pk=pk)
        if lease_req.status != 'confirmed':
            messages.error(request, 'Можно оформить договор только по подтверждённой заявке.')
            return redirect('chat:thread', request_id=pk)
        form = ContractFromRequestForm(request.POST, lease_request=lease_req)
        if form.is_valid():
            contract = form.save(commit=False)
            contract.equipment = lease_req.equipment
            contract.created_by = request.current_account
            contract.lease_request = lease_req
            contract.save()
            # Привязываем компанию к клиенту, если ещё не привязана
            if not contract.company.account_id:
                contract.company.account = lease_req.account
                contract.company.save(update_fields=['account'])
            messages.success(request, f'Договор {contract.contract_number} создан. Техника появится в «Моя техника» у клиента.')
            return redirect('chat:thread', request_id=pk)
        return render(request, 'manager/lease_request_create_contract.html', {
            'lease_request': lease_req,
            'form': form,
        })


class ChatView(ManagerRequiredMixin, View):
    """Чат с клиентами — список заявок с переходом в чат."""

    def get(self, request):
        from apps.leasing.models import LeaseRequest
        lease_requests = LeaseRequest.objects.select_related(
            'equipment', 'account', 'account__profile'
        ).order_by('-created_at')
        return render(request, 'manager/chat.html', {
            'lease_requests': lease_requests,
        })


class MaintenanceChatView(ManagerRequiredMixin, View):
    """Заявки на ТО — список с переходом в чат."""

    def get(self, request):
        maintenance_requests = MaintenanceRequest.objects.select_related(
            'equipment', 'equipment__category', 'company', 'company__account', 'company__account__profile'
        ).order_by('-created_at')
        return render(request, 'manager/maintenance_chat.html', {
            'maintenance_requests': maintenance_requests,
        })
